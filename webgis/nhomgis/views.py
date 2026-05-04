from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Product, Category, Order, OrderItem, Warehouse, ProductImage, Review, Post, StoreReview, Stock, InventoryTransaction, Comment, AboutPage, AboutSection
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum # Thư viện hỗ trợ tìm kiếm nâng cao
from django.core.exceptions import PermissionDenied
from .models import CustomerProfile
from .form import UserUpdateForm, ProfileUpdateForm
import random, uuid, json
from django.core.mail import send_mail
from django.urls import reverse
from .models import EmailVerification
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import pandas as pd
import math # Thêm thư viện math vào nếu đầu file chưa có
import requests # Đảm bảo có thư viện requests để gọi OSRM


def is_admin(user):
    if not user.is_authenticated: return False
    if user.is_staff: return True
    raise PermissionDenied


# ==========================================
# 1. PUBLIC (KHÁCH HÀNG)
# ==========================================
def home(request):
    categories = Category.objects.all()
    products_list = Product.objects.all().order_by('-id')

    for p in products_list:
        all_imgs = []
        if p.image: all_imgs.append(p.image.url) 
        for sub_img in p.images.all(): all_imgs.append(sub_img.image.url)
        if not all_imgs: all_imgs.append("https://via.placeholder.com/300")
        p.image_list = all_imgs 

    search_query, cat_id = request.GET.get('q', ''), request.GET.get('category', '')
    if search_query: products_list = products_list.filter(name__icontains=search_query)
    if cat_id and cat_id != 'all': products_list = products_list.filter(category_id=cat_id)

    page_obj = Paginator(products_list, 8).get_page(request.GET.get('page'))

    store_reviews = StoreReview.objects.all().order_by('-created_at')[:5]
    has_store_reviewed = StoreReview.objects.filter(user=request.user).exists() if request.user.is_authenticated else False

    return render(request, 'index.html', {
        'page_obj': page_obj, 'categories': categories, 'search_query': search_query, 'cat_id': cat_id,
        'store_reviews': store_reviews, 'has_store_reviewed': has_store_reviewed
    })

def product_detail(request, id):
    product = get_object_or_404(Product, id=id)
    has_reviewed = Review.objects.filter(user=request.user, product=product).exists() if request.user.is_authenticated else False
    return render(request, 'product-detail.html', {'product': product, 'reviews': product.reviews.all().order_by('-created_at'), 'has_reviewed': has_reviewed})

@login_required(login_url='/login/')
def add_review(request, product_id):
    if request.method == 'POST':
        product = get_object_or_404(Product, id=product_id)
        if not Review.objects.filter(user=request.user, product=product).exists():
            Review.objects.create(user=request.user, product=product, stars=int(request.POST.get('stars', 5)), comment=request.POST.get('comment', ''))
            messages.success(request, 'Đánh giá đã được ghi nhận!')
    return redirect('product_detail', id=product_id)

@login_required(login_url='/login/')
def add_store_review(request):
    if request.method == 'POST' and not StoreReview.objects.filter(user=request.user).exists():
        StoreReview.objects.create(user=request.user, stars=int(request.POST.get('stars', 5)), comment=request.POST.get('comment', ''))
        messages.success(request, "Cảm ơn bạn đã đánh giá cửa hàng!")
    return redirect('home')

# Sửa lại hàm blog_list trong views.py
def blog_list(request): 
    query = request.GET.get('q', '')
    posts = Post.objects.all().order_by('-created_at')
    
    if query:
        # SỬA Ở ĐÂY: Chỉ tìm kiếm trong Tiêu đề (title__icontains), bỏ nội dung đi
        posts = posts.filter(title__icontains=query)
        
    page_obj = Paginator(posts, 6).get_page(request.GET.get('page'))
    
    return render(request, 'blog.html', {
        'page_obj': page_obj, 
        'search_query': query
    })
# Cập nhật lại views.py
def blog_detail(request, id): 
    post = get_object_or_404(Post, id=id)
    # CHỈ LẤY CÁC BÌNH LUẬN GỐC (parent=None) hiển thị ở ngoài cùng
    comments = post.comments.filter(parent__isnull=True).order_by('-created_at')

    if request.method == 'POST':
        if not request.user.is_authenticated:
            messages.error(request, "Bạn cần đăng nhập để bình luận!")
            return redirect('login')
            
        content = request.POST.get('content')
        parent_id = request.POST.get('parent_id') # Lấy ID của bình luận được trả lời

        if content:
            parent_comment = None
            if parent_id: # Nếu có ID thì gán làm bình luận con
                parent_comment = Comment.objects.get(id=parent_id)

            Comment.objects.create(post=post, user=request.user, content=content, parent=parent_comment)
            messages.success(request, "Đã gửi bình luận!")
            return redirect('blog_detail', id=post.id)

    return render(request, 'blog_detail.html', {'post': post, 'comments': comments})

@login_required(login_url='/login/')
def like_comment(request, comment_id):
    comment = get_object_or_404(Comment, id=comment_id)
    if request.user in comment.likes.all():
        comment.likes.remove(request.user) # Bỏ like nếu đã like rồi
    else:
        comment.likes.add(request.user) # Thích nếu chưa
    return redirect('blog_detail', id=comment.post.id)

@user_passes_test(is_admin, login_url='/login/')
def delete_comment(request, comment_id):
    comment = get_object_or_404(Comment, id=comment_id)
    post_id = comment.post.id
    comment.delete()
    messages.success(request, "Admin đã xóa bình luận!")
    return redirect('blog_detail', id=post_id)

@login_required(login_url='/login/')
def blog_create(request):
    if not request.user.is_staff: return redirect('blog_list')
    if request.method == 'POST':
        Post.objects.create(title=request.POST.get('title'), content=request.POST.get('content'), thumbnail=request.FILES.get('thumbnail'), author=request.user)
        messages.success(request, "Đã đăng bài thành công!")
        return redirect('blog_list')
    return render(request, 'blog_form.html')

# ==========================================
# 2. TÀI KHOẢN & GIỎ HÀNG
# ==========================================
def login_view(request):
    if request.method == 'POST':
        user = authenticate(request, username=request.POST.get('username'), password=request.POST.get('password'))
        if user:
            login(request, user)
            return redirect('dashboard') if user.is_staff else redirect('home')
        return render(request, 'login.html', {'error_message': 'Sai tài khoản!'})
    return render(request, 'login.html')

def logout_view(request): logout(request); return redirect('/login/')

def register(request):
    if request.method == 'POST':
        u, e, p = request.POST.get('username'), request.POST.get('email'), request.POST.get('password')
        if User.objects.filter(username=u).exists() or User.objects.filter(email=e).exists():
            messages.error(request, "Tài khoản hoặc Email đã tồn tại!")
            return render(request, 'register.html')
        user = User.objects.create_user(username=u, email=e, password=p)
        user.is_active = False; user.save()
        otp_code = str(random.randint(100000, 999999))
        EmailVerification.objects.create(user=user, code=otp_code)
        send_mail('[KINGMATE] Mã xác nhận', f'Mã xác nhận của bạn là: {otp_code}.', 'noreply@kingmate.com', [e])
        request.session['pending_user_id'] = user.id
        return redirect('verify_email')
    return render(request, 'register.html')

def verify_email(request):
    user_id = request.session.get('pending_user_id')
    if not user_id: return redirect('register')
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        verify_obj = EmailVerification.objects.filter(user=user, code=request.POST.get('otp')).last()
        if verify_obj and not verify_obj.is_expired():
            user.is_active = True; user.save(); verify_obj.delete()
            messages.success(request, "Kích hoạt thành công!")
            return redirect('login')
        messages.error(request, "Mã sai hoặc hết hạn!")
    return render(request, 'verify_email.html')

def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        user = User.objects.filter(email=email).first()
        if user:
            otp = str(random.randint(100000, 999999))
            EmailVerification.objects.create(user=user, code=otp)
            send_mail('[KINGMATE] Khôi phục mật khẩu', f'Mã của bạn: {otp}', 'noreply@kingmate.com', [email])
            request.session['reset_email'] = email
            return redirect('verify_reset_otp')
        messages.error(request, "Email không tồn tại!")
    return render(request, 'forgot_password.html')

def verify_reset_otp(request):
    email = request.session.get('reset_email')
    if request.method == 'POST':
        user = User.objects.filter(email=email).first()
        otp_obj = EmailVerification.objects.filter(user=user, code=request.POST.get('otp')).last()
        if otp_obj and not otp_obj.is_expired():
            request.session['otp_verified'] = True
            return redirect('reset_password_new')
        messages.error(request, "Mã sai hoặc hết hạn!")
    return render(request, 'verify_reset_otp.html')

def reset_password_new(request):
    if not request.session.get('otp_verified'): return redirect('forgot_password')
    if request.method == 'POST':
        if request.POST.get('password') != request.POST.get('confirm_password'):
            messages.error(request, "Mật khẩu không khớp!"); return render(request, 'reset_password_new.html')
        user = User.objects.filter(email=request.session.get('reset_email')).first()
        if user:
            user.set_password(request.POST.get('password')); user.save()
            del request.session['reset_email']; del request.session['otp_verified']
            messages.success(request, "Đổi mật khẩu thành công!")
            return redirect('login')
    return render(request, 'reset_password_new.html')

def cart_view(request):
    cart, items, total = request.session.get('cart', {}), [], 0
    for p_id, qty in cart.items():
        try:
            p = Product.objects.get(id=p_id)
            items.append({'product': p, 'quantity': qty, 'total': p.price * qty})
            total += p.price * qty
        except: continue
    warehouses = Warehouse.objects.all()
    return render(request, 'cart.html', {'cart_items': items, 'total_price': total, 'warehouses': warehouses})

def cart_add(request):
    if request.method == 'POST':
        p_id, qty = request.POST.get('product_id'), int(request.POST.get('quantity', 1))
        cart = request.session.get('cart', {})
        cart[p_id] = cart.get(p_id, 0) + qty; request.session['cart'] = cart
    return redirect('cart')

def cart_remove(request):
    if request.method == 'POST':
        cart = request.session.get('cart', {})
        cart.pop(request.POST.get('product_id'), None); request.session['cart'] = cart
    return redirect('cart')
def cart_update(request):
    if request.method == 'POST':
        p_id = request.POST.get('product_id')
        try:
            qty = int(request.POST.get('quantity', 1))
        except ValueError:
            qty = 1
            
        cart = request.session.get('cart', {})
        if p_id in cart:
            if qty > 0:
                cart[p_id] = qty # Ghi đè bằng số lượng mới khách hàng nhập
            else:
                cart.pop(p_id, None) # Nếu nhập số <= 0 thì tự động xóa
        request.session['cart'] = cart
    return redirect('cart')
@login_required(login_url='/login/')
def checkout(request):
    if request.method == 'POST':
        lat_raw, lon_raw = request.POST.get('lat'), request.POST.get('lon')
        try: customer_lat, customer_lon = float(lat_raw) if lat_raw else None, float(lon_raw) if lon_raw else None
        except: customer_lat = customer_lon = None

        warehouse_id = request.POST.get('warehouse_id')
        warehouse = Warehouse.objects.filter(id=warehouse_id).first() if warehouse_id else None

        with transaction.atomic():
            # TẠO ĐƠN HÀNG Ở TRẠNG THÁI CHỜ - KHÔNG TRỪ KHO TẠI ĐÂY
            order = Order.objects.create(
                user=request.user, 
                shipping_address=request.POST.get('address'), 
                shipping_fee=int(float(request.POST.get('shipping_fee', 0))),
                customer_lat=customer_lat, 
                customer_lon=customer_lon,
                warehouse=warehouse,
                status='CHỜ XÁC NHẬN' # Mặc định trạng thái ban đầu
            )
            
            cart = request.session.get('cart', {})
            total = 0
            
            for p_id, qty in cart.items():
                p = Product.objects.get(id=p_id)
                OrderItem.objects.create(order=order, product=p, quantity=qty, price_at_purchase=p.price)
                total += p.price * qty
            
            order.total_amount = total + order.shipping_fee
            order.save()
            
        request.session['cart'] = {}
        messages.success(request, '🎉 Đặt hàng thành công! Đơn hàng đang chờ quản trị viên xác nhận xuất kho.')
        return redirect('home')
    return redirect('cart')

def about(request):
    about_page = AboutPage.objects.first()
    return render(request, 'about.html', {'about_page': about_page})

@login_required(login_url='/login/')
def profile_view(request): 
    # Lấy hoặc tạo profile cho user
    profile, created = CustomerProfile.objects.get_or_create(user=request.user)
    
    # Lấy toàn bộ danh sách đơn hàng của user, sắp xếp mới nhất lên đầu
    orders_list = Order.objects.filter(user=request.user).order_by('-order_date')
    
    # Cấu hình phân trang: 6 đơn hàng mỗi trang
    paginator = Paginator(orders_list, 6) 
    page_number = request.GET.get('page')
    orders = paginator.get_page(page_number)
    
    return render(request, 'profile_display.html', {
        'profile': profile,
        'orders': orders
    })

@login_required(login_url='/login/')
def profile_edit_view(request):
    profile, created = CustomerProfile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        u_form, p_form = UserUpdateForm(request.POST, instance=request.user), ProfileUpdateForm(request.POST, instance=profile)
        if u_form.is_valid() and p_form.is_valid():
            u_form.save(); p_form.save(); messages.success(request, 'Đã cập nhật thông tin!')
            return redirect('profile')
    else: u_form, p_form = UserUpdateForm(instance=request.user), ProfileUpdateForm(instance=profile)
    return render(request, 'profile_edit.html', {'u_form': u_form, 'p_form': p_form})



def api_calculate_shipping(request):
    if request.method == 'POST':
        # Lấy kho hàng được chọn
        w = get_object_or_404(Warehouse, id=request.POST.get('warehouse_id')) if request.POST.get('warehouse_id') else Warehouse.objects.first()
        if not w: 
            return JsonResponse({'error': 'Chưa có kho'}, status=400)
            
        try:
            # Gọi API lấy khoảng cách từ OSRM
            url = f"http://router.project-osrm.org/route/v1/driving/{w.longitude},{w.latitude};{request.POST.get('lng')},{request.POST.get('lat')}?overview=false"
            res = requests.get(url).json()
            
            if res.get('code') == 'Ok':
                dist = res['routes'][0]['distance'] / 1000
                
                # --- CÔNG THỨC TÍNH PHÍ VẬN CHUYỂN CHUẨN ---
                if dist <= 3.0:
                    fee = 0 # Dưới 3km: Free ship
                elif dist <= 5.0:
                    fee = w.base_fee # Từ 3km đến 5km: Lấy phí cố định (15.000)
                else:
                    # Từ 5km trở lên: Cứ quá 1km tính thêm tiền
                    # math.ceil sẽ làm tròn lên (vd: 5.2km -> vượt 0.2km -> làm tròn lên thành 1km)
                    over_km = math.ceil(dist - 5.0)
                    fee = w.base_fee + (over_km * w.fee_per_km) 
                # -------------------------------------------
                
                return JsonResponse({
                    'status': 'success', 
                    'distance': round(dist, 2), 
                    'fee': int(fee), 
                    'warehouse_name': w.name
                })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
            
    return JsonResponse({'error': 'Lỗi phương thức'}, status=400)

def shipping_page(request): return render(request, 'shipping.html', {'warehouses': Warehouse.objects.all()})

# ==========================================
# 4. ADMIN (QUẢN LÝ)
# ==========================================
@user_passes_test(is_admin, login_url='/login/')
def dashboard(request): return render(request, 'admin/dashboard.html')

@user_passes_test(is_admin, login_url='/login/')
def products_list(request): 
    query = request.GET.get('q', '')
    products = Product.objects.all().order_by('-id')
    if query:
        products = products.filter(Q(name__icontains=query) | Q(category__name__icontains=query))
    
    paginator = Paginator(products, 10)
    return render(request, 'admin/products.html', {'products': paginator.get_page(request.GET.get('page')), 'search_query': query})

@user_passes_test(is_admin, login_url='/login/')
def product_form(request, id=None): 
    product = get_object_or_404(Product, id=id) if id else None
    stock_details = []
    
    # Nếu là giao diện "Sửa Sản Phẩm" (đã có ID), tiến hành lấy chi tiết các kho
from django.db.models import Sum # Nhớ thêm dòng này ở đầu file views.py nếu chưa có

@user_passes_test(is_admin, login_url='/login/')
def product_form(request, id=None): 
    product = get_object_or_404(Product, id=id) if id else None
    stock_details = []
    
    if product:
        # --- ĐOẠN MỚI THÊM: TỰ ĐỘNG ĐỒNG BỘ DỌN RÁC ---
        # Lấy tổng số lượng thực tế từ các kho (nếu không có kho nào thì mặc định là 0)
        real_total = product.stocks.aggregate(total=Sum('quantity'))['total'] or 0
        
        # Nếu số ảo đang lưu khác với số thực tế -> Ép cập nhật lại cho đúng
        if product.stock_quantity != real_total:
            product.stock_quantity = real_total
            product.save()
        # ---------------------------------------------
            
        stock_details = Stock.objects.filter(product=product).select_related('warehouse')

    return render(request, 'admin/product-form.html', {
        'categories': Category.objects.all(), 
        'product': product,
        'stock_details': stock_details
    })

@user_passes_test(is_admin, login_url='/login/')
def product_save(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        price_raw = request.POST.get('price', '0').replace('.', '')
        unit = request.POST.get('unit')
        cat = Category.objects.get(id=request.POST.get('category_id'))
        desc = request.POST.get('description', '')
        prod_id = request.POST.get('id')
        main_img = request.FILES.get('image')
        gallery = request.FILES.getlist('gallery_images')
        
        try: 
            price = int(price_raw)
        except: 
            price = 0

        if prod_id:
            # TRƯỜNG HỢP: Cập nhật sản phẩm
            p = Product.objects.get(id=prod_id)
            p.name = name
            p.price = price
            p.unit = unit
            p.category = cat
            p.description = desc
            if main_img: 
                p.image = main_img
            # KHÔNG cập nhật p.stock_quantity ở đây để tránh đè dữ liệu kho
            p.save()
        else: 
            # TRƯỜNG HỢP: Thêm sản phẩm mới
            # Mặc định tồn kho = 0. Muốn có hàng phải vào chức năng "Nhập Kho"
            p = Product.objects.create(
                name=name, 
                price=price, 
                unit=unit, 
                stock_quantity=0, # Gán cứng bằng 0
                category=cat, 
                image=main_img, 
                description=desc
            )
            
        for img in gallery: 
            ProductImage.objects.create(product=p, image=img)
            
    return redirect('products')
@user_passes_test(is_admin, login_url='/login/')
def product_delete(request, id): get_object_or_404(Product, id=id).delete(); return redirect('products')

@user_passes_test(is_admin, login_url='/login/')
def categories_list(request): return render(request, 'admin/categories.html', {'categories': Category.objects.all()})
@user_passes_test(is_admin, login_url='/login/')
def category_form(request): return render(request, 'admin/category-form.html')
@user_passes_test(is_admin, login_url='/login/')
def category_save(request):
    if request.method == 'POST': Category.objects.create(name=request.POST.get('name'))
    return redirect('categories')
@user_passes_test(is_admin, login_url='/login/')
def category_delete(request, id): get_object_or_404(Category, id=id).delete(); return redirect('categories')

@user_passes_test(is_admin, login_url='/login/')
def warehouse_list(request): return render(request, 'admin/warehouses.html', {'warehouses': Warehouse.objects.all()})

@user_passes_test(is_admin, login_url='/login/')
def orders_list(request): 
    query = request.GET.get('q', '')
    orders = Order.objects.all().order_by('-order_date')
    if query:
        orders = orders.filter(Q(id__icontains=query) | Q(user__username__icontains=query) | Q(status__icontains=query))
    
    paginator = Paginator(orders, 10)
    return render(request, 'admin/orders.html', {'orders': paginator.get_page(request.GET.get('page')), 'search_query': query})

@user_passes_test(is_admin, login_url='/login/')
def order_update_status(request):
    if request.method == 'POST':
        order_id = request.POST.get('id')
        new_status = request.POST.get('status')
        order = get_object_or_404(Order, id=order_id)
        
        old_status = order.status
        if old_status == new_status:
            return redirect('orders')

        with transaction.atomic():
            current_batch_id = str(uuid.uuid4())[:8].upper()

            # --- TRƯỜNG HỢP: CHỐT ĐƠN / ĐÃ GIAO (BẮT ĐẦU XUẤT KHO) ---
            if new_status == 'ĐÃ GIAO':
                # Chỉ trừ kho nếu đơn hàng này chưa từng ở trạng thái ĐÃ GIAO (tránh trừ trùng)
                if old_status != 'ĐÃ GIAO':
                    for item in order.items.all():
                        if order.warehouse:
                            # 1. Trừ kho chi tiết (Warehouse Stock)
                            stock_record, _ = Stock.objects.get_or_create(warehouse=order.warehouse, product=item.product)
                            stock_record.quantity -= item.quantity
                            stock_record.save()
                            
                            # 2. Ghi nhật ký xuất hóa đơn
                            InventoryTransaction.objects.create(
                                warehouse=order.warehouse, product=item.product,
                                transaction_type='EXPORT', quantity=item.quantity,
                                note=f"Xuất kho hóa đơn đơn hàng #{order.id}",
                                user=request.user, batch_id=current_batch_id
                            )
                        
                        # 3. Cập nhật tổng tồn kho hiển thị (Model Product)
                        item.product.stock_quantity -= item.quantity
                        item.product.save()

            # --- TRƯỜNG HỢP: HỦY ĐƠN HÀNG (XỬ LÝ HOÀN KHO) ---
            elif new_status == 'ĐÃ HỦY':
                # Chỉ hoàn kho nếu trạng thái CŨ là ĐÃ GIAO (vì lúc đó hàng mới bị trừ)
                if old_status == 'ĐÃ GIAO':
                    for item in order.items.all():
                        if order.warehouse:
                            # 1. Cộng lại kho chi tiết
                            stock_record, _ = Stock.objects.get_or_create(warehouse=order.warehouse, product=item.product)
                            stock_record.quantity += item.quantity
                            stock_record.save()
                            
                            # 2. Ghi nhật ký nhập hoàn trả
                            InventoryTransaction.objects.create(
                                warehouse=order.warehouse, product=item.product,
                                transaction_type='IMPORT', quantity=item.quantity,
                                note=f"Hoàn kho từ đơn hủy #{order.id}",
                                user=request.user, batch_id=current_batch_id
                            )
                        
                        # 3. Cộng lại tổng tồn kho Product
                        item.product.stock_quantity += item.quantity
                        item.product.save()

            # Cập nhật trạng thái cuối cùng vào database
            order.status = new_status
            order.save()
            messages.success(request, f"Đã cập nhật đơn hàng #{order.id} sang {new_status} và xử lý kho tương ứng.")

    return redirect('orders')

@user_passes_test(is_admin, login_url='/login/')
def users_list(request): 
    query = request.GET.get('q', '')
    users = User.objects.filter(is_superuser=False).order_by('-id')
    if query:
        users = users.filter(Q(username__icontains=query) | Q(email__icontains=query))
    
    paginator = Paginator(users, 10)
    return render(request, 'admin/users.html', {'users': paginator.get_page(request.GET.get('page')), 'search_query': query})

@user_passes_test(is_admin, login_url='/login/')
def user_delete(request, id): get_object_or_404(User, id=id).delete(); return redirect('users')

@user_passes_test(is_admin, login_url='/login/')
def admin_map_view(request): return render(request, 'admin/map_clustering.html')

@user_passes_test(is_admin, login_url='/login/')
def api_orders_locations(request):
    orders = Order.objects.exclude(status__in=['ĐÃ GIAO', 'ĐÃ HỦY']).exclude(customer_lat__isnull=True)
    orders_data = [{'id': o.id, 'customer': o.user.get_full_name() or o.user.username, 'lat': o.customer_lat, 'lon': o.customer_lon, 'status': o.status, 'total': o.total_amount, 'address': o.shipping_address} for o in orders]
    
    warehouses = Warehouse.objects.all()
    warehouses_data = [{'id': w.id, 'name': w.name, 'lat': w.latitude, 'lon': w.longitude, 'address': w.address} for w in warehouses]
    
    return JsonResponse({'orders': orders_data, 'warehouses': warehouses_data})

import uuid # Thêm ở đầu file để tạo mã ngẫu nhiên

@user_passes_test(is_admin, login_url='/login/')
def inventory_manage(request):
    warehouses, products = Warehouse.objects.all(), Product.objects.all()
    transactions = Paginator(InventoryTransaction.objects.all().order_by('-date', '-id'), 10).get_page(request.GET.get('page'))

    stock_dict = {}
    for w in warehouses:
        stock_dict[w.id] = {}
        for s in w.stocks.all():
            stock_dict[w.id][s.product.id] = s.quantity
    stock_data_json = json.dumps(stock_dict)

    if request.method == 'POST':
        warehouse = get_object_or_404(Warehouse, id=request.POST.get('warehouse'))
        transaction_type, note = request.POST.get('transaction_type'), request.POST.get('note', '')

        # TẠO MÃ PHIẾU CHUNG CHO LẦN BẤM NÀY
        current_batch_id = str(uuid.uuid4())[:8].upper() 

        try:
            if 'excel_file' in request.FILES and request.FILES['excel_file']:
                import pandas as pd
                try:
                    df = pd.read_excel(request.FILES['excel_file'])
                    
                    with transaction.atomic():
                        for index, row in df.iterrows():
                            # --- ĐÃ FIX: KIỂM TRA VÀ BỎ QUA DÒNG TRỐNG / TỔNG CỘNG ---
                            if pd.isna(row.get('Vật Liệu')) or pd.isna(row.get('Số Lượng')):
                                continue
                                
                            if str(row.get('Vật Liệu', '')).strip().upper() == 'TỔNG CỘNG:':
                                continue

                            product_name = str(row['Vật Liệu']).strip()
                            
                            # Ép kiểu an toàn (nếu ô số lượng có chữ linh tinh thì bỏ qua dòng đó luôn)
                            try:
                                quantity = int(row['Số Lượng'])
                            except (ValueError, TypeError):
                                continue 
                            # ---------------------------------------------------------
                            
                            if quantity <= 0: continue

                            # Tìm sản phẩm theo tên (vì file xuất không có ID sản phẩm)
                            product = get_object_or_404(Product, name=product_name)
                            stock_record, _ = Stock.objects.get_or_create(warehouse=warehouse, product=product)

                            if transaction_type == 'IMPORT':
                                stock_record.quantity += quantity
                                product.stock_quantity += quantity 
                            elif transaction_type == 'EXPORT':
                                if stock_record.quantity >= quantity:
                                    stock_record.quantity -= quantity
                                    product.stock_quantity -= quantity
                                else:
                                    raise ValueError(f"Sản phẩm {product_name} không đủ hàng!")
                            
                            stock_record.save()
                            product.save()

                            InventoryTransaction.objects.create(
                                warehouse=warehouse, product=product, 
                                transaction_type=transaction_type, quantity=quantity, 
                                note="Nhập từ file Excel: " + note, user=request.user,
                                batch_id=current_batch_id
                            )
                    messages.success(request, 'Nhập dữ liệu từ Excel thành công!')
                except Exception as e:
                    messages.error(request, f"Lỗi file Excel: {str(e)}")
            else:
                p_ids, qtys = request.POST.getlist('product'), request.POST.getlist('quantity')
                with transaction.atomic():
                    for p_id, qty_str in zip(p_ids, qtys):
                        quantity = int(qty_str)
                        product = get_object_or_404(Product, id=p_id)
                        stock_record, _ = Stock.objects.get_or_create(warehouse=warehouse, product=product)

                        if transaction_type == 'IMPORT': stock_record.quantity += quantity; product.stock_quantity += quantity 
                        elif transaction_type == 'EXPORT':
                            if stock_record.quantity >= quantity: stock_record.quantity -= quantity; product.stock_quantity -= quantity
                            else: raise ValueError(f"Thiếu hàng!")
                        stock_record.save(); product.save()
                        
                        # THÊM batch_id KHI LƯU
                        InventoryTransaction.objects.create(warehouse=warehouse, product=product, transaction_type=transaction_type, quantity=quantity, note=note, user=request.user, batch_id=current_batch_id)
                messages.success(request, 'Giao dịch thành công!')
        except Exception as e: messages.error(request, str(e))
        return redirect('inventory_manage')

    return render(request, 'admin/inventory_form.html', {
        'warehouses': warehouses, 'products': products, 'transactions': transactions, 'stock_data_json': stock_data_json 
    })

@user_passes_test(is_admin, login_url='/login/')
def print_inventory_receipt(request, transaction_id):
    # Lấy bản ghi được click
    t_root = get_object_or_404(InventoryTransaction, id=transaction_id)
    
    # Lấy tất cả sản phẩm có chung batch_id
    if t_root.batch_id:
        all_items = InventoryTransaction.objects.filter(batch_id=t_root.batch_id)
    else:
        all_items = InventoryTransaction.objects.filter(id=transaction_id)
    
    # --- ĐOẠN THÊM MỚI BẮT ĐẦU TỪ ĐÂY ---
    total_amount = 0
    # Tính toán thành tiền cho từng sản phẩm và cộng dồn vào tổng tiền phiếu
    for item in all_items:
        # Gán thêm thuộc tính 'total_price' tạm thời cho từng item (Số lượng * Giá gốc của sản phẩm)
        item.total_price = item.quantity * item.product.price
        # Cộng dồn vào tổng
        total_amount += item.total_price
    # --- KẾT THÚC ĐOẠN THÊM MỚI ---

    return render(request, 'admin/inventory_receipt.html', {
        'transaction': t_root,
        'all_items': all_items, # Truyền danh sách vào template
        'total_amount': total_amount
    })

@user_passes_test(is_admin, login_url='/login/')
def export_inventory_excel(request, transaction_id):
    # 1. Lấy dữ liệu phiếu
    t_root = get_object_or_404(InventoryTransaction, id=transaction_id)
    
    if t_root.batch_id:
        items = InventoryTransaction.objects.filter(batch_id=t_root.batch_id).order_by('id')
        filename = f"Phieu_{t_root.batch_id}.xlsx"
    else:
        items = InventoryTransaction.objects.filter(id=transaction_id)
        filename = f"Phieu_Le_{t_root.id}.xlsx"

    # 2. Khởi tạo Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Chi tiết phiếu"

    # 3. Định nghĩa Header (Thêm cột Đơn Giá và Thành Tiền)
    headers = ['STT', 'Vật Liệu', 'Đơn Vị', 'Số Lượng', 'Đơn Giá (đ)', 'Thành Tiền (đ)', 'Kho', 'Loại', 'Thời Gian']
    ws.append(headers)
    
    # Định dạng Header: In đậm và căn giữa
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 4. Đổ dữ liệu và tính toán
    grand_total = 0 # Biến để tính tổng cộng cả phiếu
    
    for index, item in enumerate(items, start=1):
        unit_price = item.product.price
        line_total = item.quantity * unit_price
        grand_total += line_total
        
        ws.append([
            index,
            item.product.name,
            item.product.unit,
            item.quantity,
            unit_price,
            line_total,
            item.warehouse.name,
            "Nhập" if item.transaction_type == 'IMPORT' else "Xuất",
            item.date.strftime("%d/%m/%Y %H:%M")
        ])

    # 5. Thêm dòng Tổng Cộng ở cuối bảng
    # Chúng ta để trống các cột đầu, ghi chữ "Tổng cộng" ở cột Số lượng và giá trị ở cột Thành tiền
    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=5).value = "TỔNG CỘNG:"
    ws.cell(row=last_row, column=5).font = Font(bold=True)
    
    ws.cell(row=last_row, column=6).value = grand_total
    ws.cell(row=last_row, column=6).font = Font(bold=True, color="FF0000") # Màu đỏ cho nổi bật

    # 6. Trả về file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@user_passes_test(is_admin, login_url='/login/')
def warehouse_form(request, id=None): return render(request, 'admin/warehouse_form.html', {'warehouse': get_object_or_404(Warehouse, id=id) if id else None})

@user_passes_test(is_admin, login_url='/login/')
def warehouse_save(request):
    if request.method == 'POST':
        try:
            w_id, name, address = request.POST.get('id'), request.POST.get('name'), request.POST.get('address')
            lat, lon, base_fee, fee_per_km = float(request.POST.get('latitude', 0)), float(request.POST.get('longitude', 0)), int(request.POST.get('base_fee', 0)), int(request.POST.get('fee_per_km', 0))
            if w_id:
                w = Warehouse.objects.get(id=w_id)
                w.name, w.address, w.latitude, w.longitude, w.base_fee, w.fee_per_km = name, address, lat, lon, base_fee, fee_per_km; w.save()
                messages.success(request, 'Cập nhật kho thành công!')
            else:
                Warehouse.objects.create(name=name, address=address, latitude=lat, longitude=lon, base_fee=base_fee, fee_per_km=fee_per_km)
                messages.success(request, 'Thêm kho mới thành công!')
        except ValueError: messages.error(request, 'Lỗi định dạng số!')
    return redirect('warehouses')


@user_passes_test(is_admin, login_url='/login/')
def admin_blogs(request):
    query = request.GET.get('q', '')
    posts = Post.objects.all().order_by('-created_at')
    if query:
        posts = posts.filter(title__icontains=query)
    
    paginator = Paginator(posts, 10)
    return render(request, 'admin/blogs.html', {'posts': paginator.get_page(request.GET.get('page')), 'search_query': query})

@user_passes_test(is_admin, login_url='/login/')
def admin_blog_form(request, id=None):
    # Dùng chung 1 form cho cả Thêm và Sửa
    post = get_object_or_404(Post, id=id) if id else None
    return render(request, 'blog_form.html', {'post': post})

@user_passes_test(is_admin, login_url='/login/')
def admin_blog_save(request):
    if request.method == 'POST':
        post_id = request.POST.get('id')
        title = request.POST.get('title')
        content = request.POST.get('content')
        thumbnail = request.FILES.get('thumbnail')

        if post_id: # NẾU CÓ ID -> SỬA
            post = Post.objects.get(id=post_id)
            post.title = title
            post.content = content
            if thumbnail: 
                post.thumbnail = thumbnail
            post.save()
            messages.success(request, 'Đã cập nhật bài viết thành công!')
        else: # NẾU KHÔNG CÓ ID -> THÊM MỚI
            Post.objects.create(title=title, content=content, thumbnail=thumbnail, author=request.user)
            messages.success(request, 'Đã đăng bài viết mới!')
            
    return redirect('admin_blogs')

@user_passes_test(is_admin, login_url='/login/')
def admin_blog_delete(request, id):
    get_object_or_404(Post, id=id).delete()
    messages.success(request, 'Đã xóa bài viết!')
    return redirect('admin_blogs')

# Import thêm transaction ở đầu file nếu chưa có (from django.db import transaction)

@user_passes_test(is_admin, login_url='/login/')
def admin_about_manage(request):
    # Lấy trang AboutPage hiện tại, nếu chưa có thì tạo mới (vì chỉ cần 1 trang duy nhất)
    about_page, created = AboutPage.objects.get_or_create(id=1)
    sections = about_page.sections.all().order_by('order')

    if request.method == 'POST':
        main_title = request.POST.get('title', 'Giới thiệu về KINGMATE')
        about_page.title = main_title
        about_page.save()

        # Lấy danh sách các hộp được gửi lên từ Form (Mảng Dữ liệu)
        headings = request.POST.getlist('headings[]')
        contents = request.POST.getlist('contents[]')
        types = request.POST.getlist('types[]') 
        icons = request.POST.getlist('icons[]') 
        aligns = request.POST.getlist('aligns[]') # THÊM DÒNG NÀY

        with transaction.atomic():
            about_page.sections.all().delete()
            # THÊM BIẾN a VÀ aligns VÀO VÒNG LẶP ZIP
            for index, (h, c, t, i, a) in enumerate(zip(headings, contents, types, icons, aligns)):
                AboutSection.objects.create(
                    page=about_page,
                    heading=h,
                    content=c,
                    section_type=t,
                    icon_class=i,
                    text_align=a, # LƯU THÊM VÀO DATABASE NÀY
                    order=index
                )
        
        messages.success(request, 'Đã cập nhật trang Giới thiệu thành công!')
        return redirect('admin_about_manage')

    return render(request, 'admin/about_manage.html', {
        'about_page': about_page,
        'sections': sections
    })

# Đừng quên sửa lại hàm about (cho khách hàng xem) ở phần 1. PUBLIC:

from django.db.models import Sum
from django.http import HttpResponse

@user_passes_test(is_admin, login_url='/login/')
def clean_stock_data(request):
    products = Product.objects.all()
    fixed_count = 0
    
    for p in products:
        real_total = p.stocks.aggregate(total=Sum('quantity'))['total'] or 0
        if p.stock_quantity != real_total:
            p.stock_quantity = real_total
            p.save()
            fixed_count += 1
            
    return HttpResponse(f"<h1>Đã dọn dẹp thành công {fixed_count} sản phẩm bị dính số ảo!</h1><br><a href='/'>Quay lại trang chủ</a>")

@user_passes_test(is_admin, login_url='/login/')
def warehouse_delete(request, id):
    warehouse = get_object_or_404(Warehouse, id=id)
    
    # Kiểm tra: Nếu kho còn hàng tồn (Stock > 0) thì không cho xóa
    has_stock = Stock.objects.filter(warehouse=warehouse, quantity__gt=0).exists()
    
    if has_stock:
        messages.error(request, f"Không thể xóa kho '{warehouse.name}' vì vẫn còn hàng tồn kho trong hệ thống!")
    else:
        warehouse.delete()
        messages.success(request, f"Đã xóa kho hàng '{warehouse.name}' thành công.")
        
    return redirect('warehouses')

def custom_catch_all_404(request, *args, **kwargs):
    return render(request, '404.html', status=404)